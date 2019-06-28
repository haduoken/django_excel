import pandas as pd
from pandas import DataFrame
import os
import sys
from openpyxl import load_workbook

def get_visible_names(file):
	wb = load_workbook(file)
	visible_sheets = [sheet for sheet in wb.sheetnames if wb[sheet].sheet_state == 'visible']
	return visible_sheets

# 获取保留的属性
def get_keys(file,sheet_name):
	# out_file = os.path.join('excels','out.xlsx')
	# out_sheet_name = '明细汇总'
	# out_sheet_name = get_visible_names(out_file)
	out_frame = pd.read_excel(file,sheet_name=sheet_name)
	keys = out_frame.keys().str
	keys1 = out_frame.keys().values
	keys_out = keys1.tolist()
	return keys_out

def write_file_by_keys(keys):
	# load other files
	files = os.listdir('excels')
	f_list=[]
	out_frame = DataFrame()
	for file in files:
		if file=='out.xlsx':
			continue
		file_name = os.path.join('excels', file)
		sheet_name = get_visible_names(file_name)
		if isinstance(sheet_name,list):
			# f = pd.read_excel(file_name,sheet_name=sheet_name)
			for sheet in sheet_name:
				f = pd.read_excel(file_name,sheet_name=sheet)
				keys_have = f.keys().values.tolist()
				keys_use = [key for key in keys_out if key in keys_have]
				tmp = f[keys_use]
				out_frame = pd.concat([out_frame,tmp],sort=False)
	out_frame.to_excel('output.xlsx')
